import sys
import argparse
import openpyxl

header = '''library ieee;
use ieee.math_real.all;

package dlx_package is

	constant c_DLX_PC_WIDTH : integer := 10;
	constant c_DLX_WORD_WIDTH : integer := 32;
	constant c_NUM_OF_REGISTERS : integer := 32;
	constant c_DLX_REG_ADDR_WIDTH : integer := integer(ceil(log2(real(c_NUM_OF_REGISTERS))));
	constant c_DLX_IMM_WIDTH : integer := 16;
	constant c_DLX_OPCODE_WIDTH : integer := 6;'''

footer = '''end package dlx_package;

package body dlx_package is

end package body dlx_package;'''

INSTRUCTIONS = []

OP_CODES_DICT = {}

_OP_CODES = []

NO_OPS = []

MEMORY_OPS = []

BRANCH_OPS = []

JUMP_OPS = []

REGISTER_OPS = []

IMMEDIATE_OPS = []

_wb_obj = openpyxl.load_workbook('Instruction_Set.xlsx')
_sheet = _wb_obj.active

_row_count = 1

for row in _sheet.iter_rows(2, _sheet.max_row):
	_row_count += 1
	inst = row[0].value
	op_code = int(row[4].value, 16)
	typ = row[5].value
	
	if inst not in INSTRUCTIONS:
		INSTRUCTIONS.append(inst)
	else:
		print(f"Duplicate Instruction Found In Instruction Set Document at row {_row_count}")
		sys.exit()
	if op_code not in _OP_CODES:
		_OP_CODES.append(op_code)
		OP_CODES_DICT[inst] = op_code
	else:
		print(f"Duplicate OP-Code Found In Instruction Set Document at row {_row_count}")
		sys.exit()
	if typ == 'Register':
		REGISTER_OPS.append(inst)
	elif typ == 'Immediate':
		IMMEDIATE_OPS.append(inst)
	elif typ == 'Jump':
		JUMP_OPS.append(inst)
	elif typ == 'Branch':
		BRANCH_OPS.append(inst)
	elif typ == 'Memory':
		MEMORY_OPS.append(inst)
	elif typ == 'NOP':
		NO_OPS.append(inst)
	else:
		print(f"Unknown Instrucion Type Found In Instruction Set Document at row {_row_count}")
		sys.exit()

_wb_obj.close()

def build_python_register_file():
	with open('dlx_instruction_set.py', 'w') as outfile:

		outfile.write('\nINSTRUCTIONS = [\n')

		for inst in INSTRUCTIONS:
			outfile.write(f'\t\t"{inst}",\n')
		outfile.write(']\n\n')

		outfile.write('NO_OPS = [\n')

		for inst in NO_OPS:
			outfile.write(f'\t\t"{inst}",\n')
		outfile.write(']\n\n')

		outfile.write('MEMORY_OPS = [\n')

		for inst in MEMORY_OPS:
			outfile.write(f'\t\t"{inst}",\n')
		outfile.write(']\n\n')

		outfile.write('BRANCH_OPS = [\n')

		for inst in BRANCH_OPS:
			outfile.write(f'\t\t"{inst}",\n')
		outfile.write(']\n\n')

		outfile.write('JUMP_OPS = [\n')

		for inst in JUMP_OPS:
			outfile.write(f'\t\t"{inst}",\n')
		outfile.write(']\n\n')

		outfile.write('REGISTER_OPS = [\n')

		for inst in REGISTER_OPS:
			outfile.write(f'\t\t"{inst}",\n')
		outfile.write(']\n\n')

		outfile.write('IMMEDIATE_OPS = [\n')

		for inst in IMMEDIATE_OPS:
			outfile.write(f'\t\t"{inst}",\n')
		outfile.write(']\n\n')

		outfile.write('OP_CODES_DICT = {\n')

		for key in OP_CODES_DICT.keys():
			outfile.write(f'\t\t"{key}":{OP_CODES_DICT[key]},\n')
		outfile.write('}\n\n')

def build_dlx_package_vhd_file():
	with open('DLX_Package.vhd', 'w') as outfile:
		
		outfile.write(header)
		outfile.write('\n\n\t-- Instructions\n\n')

		for inst in OP_CODES_DICT.keys():
			op_code = OP_CODES_DICT[inst]
			if len(inst) >= 5:
				inst += '\t'
			else:
				inst += '\t\t'

			line = f'\tconstant c_DLX_{inst}: std_logic_vector(6 downto 0) := "{format(op_code, "b").zfill(6)}";\n'
			outfile.write(line)

		outfile.write('\n')
		outfile.write(footer)


if __name__ == '__main__':
	parser = argparse.ArgumentParser()
	parser.add_argument('--vhdl', action='store_true')
	parser.add_argument('--py', action='store_true')
	argument = parser.parse_args()

	build_vhdl_flag = argument.vhdl
	build_py_flag = argument.py

	if build_py_flag:
		print("Generating dlx_instruction_set.py ....")
		build_python_register_file()

	if build_vhdl_flag:
		print("Generating DLX_Package.vhd ....")
		build_dlx_package_vhd_file()
